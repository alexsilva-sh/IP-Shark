"""Planilha entregue ao cliente: o veredito precisa chegar la, com a linha destacada.

Gera .xlsx de verdade numa pasta temporaria e le de volta. Sem rede.
"""
import os
import shutil
import tempfile

from _comum import check, encerrar

import i18n
from core import api as core
from core import reputacao
from i18n import t
from openpyxl import load_workbook
from services import exportacao
from ui import apresentacao

saida = tempfile.mkdtemp(prefix="ipshark-testes-")
exportacao.escolher_diretorio = lambda parent=None, titulo=None: saida

ABUSE_RUIM = {"data": {"abuseConfidenceScore": 92, "isWhitelisted": False}}
ABUSE_WL = {"data": {"abuseConfidenceScore": 0, "isWhitelisted": True}}
VT_LIMPO = {"data": {"attributes": {"last_analysis_stats": {"malicious": 0}}}}
VT_RUIM = {"data": {"attributes": {"last_analysis_stats": {"malicious": 34}}}}


def ip(endereco, abuse, vt, **kw):
    return reputacao.build_ip_result(endereco, abuse, vt, None, "Cidade", "Pais", "d", **kw)


limpo = ip("1.1.1.1", {"data": {"abuseConfidenceScore": 0}}, VT_LIMPO, estado_ibm=None)
malicioso = ip("9.8.7.6", ABUSE_RUIM, VT_RUIM, estado_ibm=None)
revisar = ip("5.6.7.8", ABUSE_WL, VT_RUIM, estado_ibm=None)
incompleto = ip("4.4.4.4", None, VT_LIMPO, estado_abuse=core.FONTE_COTA, estado_ibm=None)

print("\n[1] O veredito chega na planilha de IP")
headers = [t("csv_ip"), t("csv_verdict"), t("csv_abuse_score"), t("csv_vt_score"),
           t("csv_md_score"),
           t("csv_domain"), t("csv_country"), t("csv_city"), t("csv_last_report"),
           t("csv_abuse_link"), t("csv_vt_link"), t("csv_md_link")]
linhas = [apresentacao.linha_planilha_ip(d, com_ibm=False)
          for d in (limpo, malicioso, revisar, incompleto)]
check(all(len(linha) == len(headers) for linha in linhas),
      f"cada linha bate com o cabecalho ({len(linhas[0])} x {len(headers)})")

caminho = exportacao.salvar_planilha(linhas, headers, filename="ip.xlsx", coluna_veredito=2)
ws = load_workbook(caminho).active
check(ws.cell(1, 2).value == t("csv_verdict"), "coluna 2 e o veredito")
check(ws.cell(2, 2).value == t("verdict_clean"), "IP limpo exportado como limpo")
check(ws.cell(3, 2).value == t("verdict_bad"), "IP malicioso exportado como malicioso")
check(ws.cell(4, 2).value == t("verdict_review"),
      "whitelist com deteccao chega como Revisar -- era o que nao saia do relatorio")
check(ws.cell(5, 2).value == t("verdict_incomplete"), "fonte em cota exportada como incompleta")
check(ws.cell(2, 1).value == "1.1.1.1", "o indicador continua na primeira coluna")

print("\n[2] Linha com achado sai destacada")
fundo = lambda linha: (ws.cell(linha, 1).fill.start_color.rgb or "")[-6:]
check(fundo(3) == "FFD9D9", f"linha maliciosa em vermelho claro ({fundo(3)})")
check(fundo(4) == "FFF2CC", f"linha de revisao em ambar ({fundo(4)})")
check(fundo(5) == "FFF2CC", "analise incompleta tambem chama atencao")
check(fundo(2) not in ("FFD9D9", "FFF2CC"), "linha limpa fica no zebrado normal")
check(ws.cell(3, 2).font.bold and ws.cell(3, 2).font.color.rgb[-6:] == "9C0006",
      "texto do veredito malicioso em negrito vermelho")
check(ws.cell(2, 2).font.color.rgb[-6:] == "1E7B34", "veredito limpo em verde")

print("\n[3] Sem coluna de veredito, o zebrado nao muda")
caminho = exportacao.salvar_planilha(linhas, headers, filename="sem_cor.xlsx")
ws2 = load_workbook(caminho).active
cores = {(ws2.cell(i, 1).fill.start_color.rgb or "")[-6:] for i in range(2, 6)}
check(cores <= {"F2F2F2", "FFFFFF"}, f"nenhuma linha destacada sem o parametro ({cores})")

print("\n[4] Hash e dominio seguem o mesmo formato")
VT_HASH = {"data": {"attributes": {"last_analysis_stats": {"malicious": 7},
                                   "meaningful_name": "fatura.exe"}}}
OTX_ZERO = core._otx_contexto({})
d_hash = reputacao.build_hash_result("a" * 32, VT_HASH, "-", OTX_ZERO)
h_headers = [t("csv_hash"), t("csv_verdict"), t("csv_vt_score"), t("csv_alien_score"),
             t("csv_md_score"),
             t("csv_file_name"), t("csv_last_analysis"), t("csv_vt_link"),
             t("csv_alien_link"), t("csv_md_link"), t("csv_joe_link")]
h_linha = apresentacao.linha_planilha_hash(d_hash, com_ibm=False)
check(len(h_linha) == len(h_headers), f"linha de hash bate com o cabecalho ({len(h_linha)})")
caminho = exportacao.salvar_planilha([h_linha], h_headers, filename="hash.xlsx",
                                     coluna_veredito=2)
ws3 = load_workbook(caminho).active
check(ws3.cell(2, 2).value == t("verdict_bad"), "veredito do hash na planilha")

d_url = reputacao.build_url_result("exemplo.com", 4, "-", OTX_ZERO)
u_headers = [t("csv_domain"), t("csv_verdict"), t("csv_vt_score"), t("csv_alien_score"),
             t("csv_md_score"), t("csv_vt_link"), t("csv_alien_link"), t("csv_md_link")]
u_linha = apresentacao.linha_planilha_url(d_url, com_ibm=False)
check(len(u_linha) == len(u_headers), f"linha de dominio bate com o cabecalho ({len(u_linha)})")

print("\n[5] A aba de IPs associados de cada dominio tambem leva o veredito")
caminho = exportacao.salvar_planilha_dominios(
    domain_results=[u_linha], domain_headers=u_headers,
    ip_results_by_domain={"exemplo.com": [apresentacao.linha_planilha_ip(malicioso, False)]},
    ip_headers=headers, filename="dominios.xlsx", coluna_veredito=2)
wb = load_workbook(caminho)
check(wb["Dominios"].cell(2, 2).value == t("verdict_bad"), "veredito do dominio")
aba_ips = wb["IPs - exemplo.com"]
check(aba_ips.cell(1, 2).value == t("csv_verdict"), "aba do IP associado tem a coluna")
check(aba_ips.cell(2, 2).value == t("verdict_bad"), "veredito do IP associado")
check((aba_ips.cell(2, 1).fill.start_color.rgb or "")[-6:] == "FFD9D9",
      "IP associado malicioso tambem sai destacado")

print("\n[5b] O link do MetaDefender chega nas tres planilhas")
MD_OK = {"detectados": 3, "total": 20}
com_md = (
    ("IP", apresentacao.linha_planilha_ip(
        ip("9.8.7.6", ABUSE_RUIM, VT_RUIM, estado_ibm=None, md=MD_OK,
           estado_md=core.FONTE_OK), False), headers),
    ("hash", apresentacao.linha_planilha_hash(
        reputacao.build_hash_result("a" * 32, VT_HASH, "-", OTX_ZERO, md=MD_OK,
                                    estado_md=core.FONTE_OK), False), h_headers),
    ("dominio", apresentacao.linha_planilha_url(
        reputacao.build_url_result("exemplo.com", 4, "-", OTX_ZERO, md=MD_OK,
                                   estado_md=core.FONTE_OK), False), u_headers),
)
for nome, linha, cabecalho in com_md:
    coluna = cabecalho.index(t("csv_md_link"))
    check(len(linha) == len(cabecalho) and linha[coluna] == reputacao.LINK_MD,
          f"planilha de {nome} leva o link na coluna certa ({linha[coluna]!r})")
    placar = cabecalho.index(t("csv_md_score"))
    check(linha[placar] == "3/20", f"e o placar do multiscanning ({linha[placar]!r})")

sem_md = apresentacao.linha_planilha_hash(d_hash, False)
check(sem_md[h_headers.index(t("csv_md_link"))] == "" and len(sem_md) == len(h_headers),
      "fonte nao consultada deixa a celula vazia, sem desalinhar a planilha")

print("\n[6] O filtro do Excel cobre a coluna nova")
check(ws.auto_filter.ref.startswith("A1"), f"auto-filtro desde A1 ({ws.auto_filter.ref})")

print("\n[7] Nome das abas vem traduzido de quem chama, nao fixo em portugues")
# Nos dois idiomas, nao so no ativo: o item existe porque o cliente que recebe a planilha
# em ingles via as abas em portugues.
for chave in ("csv_sheet_results", "csv_sheet_domains", "csv_sheet_ips_prefix"):
    faltando = [lang for lang, textos in i18n._IDIOMAS.items() if chave not in textos]
    check(not faltando, f"{chave} existe em pt e en" + (f" (falta em {faltando})" if faltando else ""))
check(i18n._IDIOMAS["en"]["csv_sheet_domains"] != i18n._IDIOMAS["pt"]["csv_sheet_domains"],
      "a aba de dominios realmente muda de idioma")

# Importar o i18n aqui dentro ligaria "t" ou "i18n" ao modulo. E a correcao errada:
# a exportacao nao pode saber em que idioma a planilha foi pedida.
check(not any(hasattr(exportacao, nome) for nome in ("t", "plural", "i18n")),
      "exportacao segue sem importar o i18n -- o nome da aba chega por parametro")

caminho = exportacao.salvar_planilha(linhas, headers, filename="traduzida.xlsx",
                                     coluna_veredito=2, aba="Results")
check(load_workbook(caminho).sheetnames == ["Results"],
      "aba da planilha de IP e de hash segue o parametro")

caminho = exportacao.salvar_planilha_dominios(
    domain_results=[u_linha], domain_headers=u_headers,
    ip_results_by_domain={"exemplo.com": [apresentacao.linha_planilha_ip(malicioso, False)]},
    ip_headers=headers, filename="dominios_traduzida.xlsx", coluna_veredito=2,
    aba="Domains", prefixo_aba_ips="IPs - ")
check(load_workbook(caminho).sheetnames == ["Domains", "IPs - exemplo.com"],
      "aba de dominio e prefixo dos IPs seguem o parametro")

# O prefixo entra na conta do teto de 31: um prefixo traduzido mais longo precisa
# cortar mais do dominio, senao o Excel recusa abrir a planilha entregue.
longo = "um-subdominio-bem-comprido.exemplo.com.br"
caminho = exportacao.salvar_planilha_dominios(
    domain_results=[u_linha], domain_headers=u_headers,
    ip_results_by_domain={longo: [apresentacao.linha_planilha_ip(malicioso, False)],
                          "outro.com:8080": [apresentacao.linha_planilha_ip(limpo, False)]},
    ip_headers=headers, filename="abas_longas.xlsx", coluna_veredito=2,
    prefixo_aba_ips="Planilha de IPs - ")
abas = load_workbook(caminho).sheetnames
check(max(len(n) for n in abas) <= exportacao.LIMITE_NOME_ABA,
      f"prefixo longo ainda cabe no teto do Excel ({max(abas, key=len)!r})")
check(all(":" not in n for n in abas), "caractere proibido no nome de aba continua trocado")

shutil.rmtree(saida, ignore_errors=True)
check(not os.path.exists(saida), "pasta temporaria removida")
encerrar()
