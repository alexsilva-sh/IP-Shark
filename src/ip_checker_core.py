import base64
import requests
import pyperclip
import ipaddress
import csv
import os
import sys
import tkinter as tk
from tkinter import filedialog
import unicodedata
import subprocess
import threading
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor
from selenium.webdriver.chrome.service import Service as ChromeService
from country_codes import COUNTRY_NAMES_LOCAL
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import secure_store

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
    BUNDLE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    BUNDLE_DIR = BASE_DIR

secure_store.migrar_se_preciso(BASE_DIR)

ABUSEIPDB_API_KEY = None
VIRUSTOTAL_API_KEY = None
IPINFO_API_KEY = None
ALIENVAULT_API_KEY = None

_lock_chaves = threading.Lock()
_mtime_store = -1
_chaves_carregadas = False

def reload_api_keys(forcar=False):
    """Recarrega as chaves do cofre, decifrando apenas quando o arquivo muda."""
    global ABUSEIPDB_API_KEY, VIRUSTOTAL_API_KEY, IPINFO_API_KEY, ALIENVAULT_API_KEY
    global _mtime_store, _chaves_carregadas
    try:
        mtime = os.path.getmtime(secure_store.caminho_store())
    except OSError:
        mtime = -1
    with _lock_chaves:
        if _chaves_carregadas and not forcar and mtime == _mtime_store:
            return
        dados = secure_store.carregar()
        _mtime_store = mtime
        _chaves_carregadas = True
        ABUSEIPDB_API_KEY = dados.get('ABUSEIPDB_API_KEY') or None
        VIRUSTOTAL_API_KEY = dados.get('VIRUSTOTAL_API_KEY') or None
        IPINFO_API_KEY = dados.get('IPINFO_API_KEY') or None
        ALIENVAULT_API_KEY = dados.get('ALIENVAULT_API_KEY') or None

def carregar_chaves_salvas():
    """Chaves atualmente no cofre, para preencher a tela de configuracao."""
    return secure_store.carregar()

def salvar_chaves(chaves):
    """Grava as chaves no cofre criptografado e recarrega o que esta em memoria."""
    caminho = secure_store.salvar(chaves)
    reload_api_keys(forcar=True)
    return caminho

reload_api_keys()

TIMEOUT_HTTP = 15

# Estado de cada fonte consultada. Indisponibilidade nunca pode virar score zero:
# sem isso, rede caida e cota estourada aparecem como "IP limpo".
FONTE_OK = "ok"                      # respondeu com dado utilizavel
FONTE_SEM_DADOS = "sem_dados"        # respondeu, mas nao conhece o indicador
FONTE_SEM_CHAVE = "sem_chave"        # chave ausente ou recusada
FONTE_COTA = "cota"                  # HTTP 429
FONTE_INDISPONIVEL = "indisponivel"  # rede, timeout, 5xx, resposta ilegivel

ESTADOS_SEM_RESPOSTA = (FONTE_SEM_CHAVE, FONTE_COTA, FONTE_INDISPONIVEL)


def _consultar(url, **kwargs):
    """GET com timeout que classifica a resposta. Devolve (json, estado)."""
    kwargs.setdefault("timeout", TIMEOUT_HTTP)
    try:
        resposta = requests.get(url, **kwargs)
    except requests.exceptions.RequestException:
        return None, FONTE_INDISPONIVEL
    if resposta.status_code == 429:
        return None, FONTE_COTA
    if resposta.status_code == 404:
        return None, FONTE_SEM_DADOS
    if resposta.status_code in (401, 403):
        return None, FONTE_SEM_CHAVE
    if resposta.status_code != 200:
        return None, FONTE_INDISPONIVEL
    try:
        return resposta.json(), FONTE_OK
    except ValueError:
        return None, FONTE_INDISPONIVEL


def classificar_ibm(valor):
    """Traduz o retorno do scraping do X-Force para um estado de fonte."""
    if valor is None:
        return None      # fonte nao consultada
    bruto = str(valor).strip().lower()
    if bruto == "error":
        return FONTE_INDISPONIVEL
    if bruto == "unknown":
        return FONTE_SEM_DADOS
    return FONTE_OK


def safe_get(d, *keys, default=None):
    for key in keys:
        if isinstance(d, dict) and key in d:
            d = d[key]
        else:
            return default
    return d

def is_valid_ip(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False

def check_ip_abuseipdb(ip):
    reload_api_keys()
    if not ABUSEIPDB_API_KEY:
        return None, FONTE_SEM_CHAVE
    return _consultar('https://api.abuseipdb.com/api/v2/check',
                      headers={'Accept': 'application/json', 'Key': ABUSEIPDB_API_KEY},
                      params={'ipAddress': ip, 'maxAgeInDays': 90})

def check_ip_virustotal(ip):
    reload_api_keys()
    if not VIRUSTOTAL_API_KEY:
        return None, FONTE_SEM_CHAVE
    return _consultar(f'https://www.virustotal.com/api/v3/ip_addresses/{ip}',
                      headers={'x-apikey': VIRUSTOTAL_API_KEY})

def check_ip_ibm(driver, ip):
    url = f"https://exchange.xforce.ibmcloud.com/ip/{ip}"
    driver.execute_script("window.open(arguments[0], '_blank');", url)
    driver.switch_to.window(driver.window_handles[-1])
    try:
        import time
        time.sleep(8)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        h1 = soup.find("h1", class_="risklevelbar")
        if h1:
            score_span = h1.find("span", class_="numtitle")
            if score_span:
                try:
                    risk_score = float(score_span.text.strip())
                except ValueError:
                    risk_score = "unknown"
            else:
                risk_class = h1.get("class", [])
                if "high" in risk_class:
                    risk_score = "high"
                elif "medium" in risk_class:
                    risk_score = "medium"
                elif "low" in risk_class:
                    risk_score = "low"
                else:
                    risk_score = "unknown"
        else:
            risk_score = "unknown"
    except Exception:
        risk_score = "error"
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    return ip, risk_score

def check_hash_ibm(driver, hash_str):
    url = f"https://exchange.xforce.ibmcloud.com/malware/{hash_str}"
    driver.execute_script("window.open(arguments[0], '_blank');", url)
    driver.switch_to.window(driver.window_handles[-1])

    try:
        WebDriverWait(driver, 18).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".risklevelbar")))
        soup = BeautifulSoup(driver.page_source, "html.parser")
        risk_element = soup.find(class_="risklevelbar")
        risk_class = risk_element.get("class", []) if risk_element else []
        if "high" in risk_class:
            score = "high"
        elif "medium" in risk_class:
            score = "medium"
        elif "low" in risk_class:
            score = "low"
        else:
            score = "unknown"
    except Exception:
        score = "error"
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    return hash_str, score

def check_hash_joesandbox(driver, hash_str):
    base_url = "https://www.joesandbox.com/analysis/search?q="
    search_url = base_url + hash_str
    driver.execute_script("window.open(arguments[0], '_blank');", search_url)
    driver.switch_to.window(driver.window_handles[-1])
    found = False
    try:
        WebDriverWait(driver, 18).until(
            EC.presence_of_element_located((By.TAG_NAME, "body")))
        found = "Full Report" in driver.page_source
    except Exception:
        pass
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    return found, search_url

def check_hash_virustotal(hash_str):
    reload_api_keys()
    if not VIRUSTOTAL_API_KEY:
        return None, FONTE_SEM_CHAVE
    return _consultar(f"https://www.virustotal.com/api/v3/files/{hash_str}",
                      headers={'x-apikey': VIRUSTOTAL_API_KEY})

def _alienvault(tipo, indicador, link):
    """Consulta o OTX. Devolve (contagem_de_pulsos, link, estado)."""
    reload_api_keys()
    if not ALIENVAULT_API_KEY:
        return None, link, FONTE_SEM_CHAVE
    dados, estado = _consultar(
        f"https://otx.alienvault.com/api/v1/indicators/{tipo}/{indicador}/general",
        headers={"X-OTX-API-KEY": ALIENVAULT_API_KEY, "Accept": "application/json"})
    if estado == FONTE_SEM_DADOS:
        return "0", link, FONTE_OK   # indicador ausente do OTX significa zero pulsos
    if estado != FONTE_OK:
        return None, link, estado
    return str(safe_get(dados, "pulse_info", "count", default=0)), link, FONTE_OK

def check_hash_alienvault(hash_str):
    return _alienvault("file", hash_str, f"https://otx.alienvault.com/indicator/file/{hash_str}")

def check_url_alienvault(url):
    return _alienvault("url", url, f"https://otx.alienvault.com/indicator/url/{url}")

def check_url_ibm(driver, url):
    ibm_url = f"https://exchange.xforce.ibmcloud.com/url/{url}"
    if len(driver.window_handles) > 1:
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
    driver.get(ibm_url)
    try:
        WebDriverWait(driver, 18).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "h2.scorebackgroundfilter.numtitle")))
        soup = BeautifulSoup(driver.page_source, "html.parser")
        elem = soup.find("h2", class_="scorebackgroundfilter numtitle")
        risk_score = elem.text.strip() if elem else "unknown"
    except Exception:
        risk_score = "error"
    return risk_score

def check_url_virustotal(url):
    reload_api_keys()
    if not VIRUSTOTAL_API_KEY:
        return {"score": None, "not_found": False}, FONTE_SEM_CHAVE
    vt_id = base64.urlsafe_b64encode(url.encode()).decode().rstrip("=")
    dados, estado = _consultar(f"https://www.virustotal.com/api/v3/urls/{vt_id}",
                               headers={"x-apikey": VIRUSTOTAL_API_KEY,
                                        "Accept": "application/json"})
    if estado == FONTE_SEM_DADOS:
        return {"score": None, "not_found": True}, FONTE_SEM_DADOS
    if estado != FONTE_OK:
        return {"score": None, "not_found": False}, estado
    score = safe_get(dados, "data", "attributes",
                     "last_analysis_stats", "malicious", default=0)
    return {"score": score, "not_found": False}, FONTE_OK
def start_browser():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--log-level=3")
    chrome_path = ChromeDriverManager().install()
    service = ChromeService(executable_path=chrome_path)
    service.creationflags = subprocess.CREATE_NO_WINDOW
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def get_location(ip):
    reload_api_keys()
    dados, estado = _consultar(f"https://ipinfo.io/{ip}/json?token={IPINFO_API_KEY}")
    if estado != FONTE_OK:
        return 'N/A', 'N/A'
    return dados.get('city', 'N/A'), translate_country_name(dados.get('country', 'N/A'))

def get_domain_from_abuseipdb(abuseipdb_result):
    try:
        return remover_acentos(abuseipdb_result['data'].get('domain', 'N/A')) if abuseipdb_result else 'N/A'
    except KeyError:
        return 'N/A'

def remover_acentos(texto):
    if isinstance(texto, str):
        return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    return texto

def is_whitelisted_abuseipdb(abuseipdb_result):
    if not abuseipdb_result:
        return False
    try:
        return bool(abuseipdb_result['data'].get('isWhitelisted', False))
    except (KeyError, TypeError, AttributeError):
        return False

def translate_country_name(country_code):
    """
    Traduz código ISO de país para nome completo
    usando dicionário local.
    """
    if not country_code or country_code == 'N/A':
        return 'N/A'

    lang = os.getenv("APP_LANG", "pt")
    code = country_code.strip().upper()

    local_dict = COUNTRY_NAMES_LOCAL.get(lang, COUNTRY_NAMES_LOCAL.get("en", {}))
    return local_dict.get(code, country_code)

def _format_worksheet(ws):
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Consolas", size=10)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    row_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.fill = fill
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 4, 60)
        ws.column_dimensions[col_letter].width = adjusted_width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

def escolher_diretorio(parent=None, titulo="Selecione a pasta para salvar"):
    """Reaproveita a janela existente. Criar um segundo tk.Tk() dentro de um app Tk
    ja em execucao gera comportamento erratico nos dialogos e vaza a raiz nova."""
    dono = parent or tk._default_root
    temporaria = None
    if dono is None:
        temporaria = tk.Tk()
        temporaria.withdraw()
        dono = temporaria
    try:
        diretorio = filedialog.askdirectory(title=titulo, parent=dono)
    finally:
        if temporaria is not None:
            temporaria.destroy()
    return diretorio if diretorio else os.getcwd()

def save_to_csv(results, headers, filename="results.xlsx", parent=None, titulo=None):
    diretorio = escolher_diretorio(parent, titulo or "Selecione a pasta para salvar")
    if filename.endswith(".csv"):
        filename = filename.replace(".csv", ".xlsx")
    filepath = os.path.join(diretorio, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(headers)
    for row in results:
        ws.append([str(v) if v is not None else "" for v in row])
    _format_worksheet(ws)
    wb.save(filepath)

def save_to_excel(domain_results, domain_headers, ip_results_by_domain, ip_headers,
                  filename="domain_results.xlsx", parent=None, titulo=None):
    diretorio = escolher_diretorio(parent, titulo or "Selecione a pasta para salvar")
    filepath = os.path.join(diretorio, filename)
    wb = Workbook()
    ws_domains = wb.active
    ws_domains.title = "Dominios"
    ws_domains.append(domain_headers)
    for row in domain_results:
        ws_domains.append([str(v) if v is not None else "" for v in row])
    _format_worksheet(ws_domains)
    for domain, ip_rows in ip_results_by_domain.items():
        safe_name = domain[:25]
        for char in ['/', '\\', '*', '?', ':', '[', ']']:
            safe_name = safe_name.replace(char, '_')
        ws_ip = wb.create_sheet(title=f"IPs - {safe_name}")
        ws_ip.append(ip_headers)
        for row in ip_rows:
            ws_ip.append([str(v) if v is not None else "" for v in row])
        _format_worksheet(ws_ip)
    wb.save(filepath)

def format_output(ip, abuseipdb_result, virustotal_result, ibm_score, city, country, domain, index):
    try:
        abuse_confidence = safe_get(abuseipdb_result, 'data', 'abuseConfidenceScore', default=0)
        vt_score = safe_get(virustotal_result, 'data', 'attributes', 'last_analysis_stats', 'malicious', default=0)
        abuseipdb_link = f"https://www.abuseipdb.com/check/{ip}"
        virustotal_link = f"https://www.virustotal.com/gui/ip-address/{ip}"
        ibm_link = f"https://exchange.xforce.ibmcloud.com/ip/{ip}" if ibm_score is not None else ""

        last_reported_at = abuseipdb_result['data'].get('lastReportedAt') if abuseipdb_result else None
        if last_reported_at:
            utc_time = datetime.fromisoformat(last_reported_at.replace('Z', '+00:00'))
            brasilia_time = utc_time - timedelta(hours=3)
            last_reported_at_formatted = brasilia_time.strftime('%d/%m/%Y %H:%M:%S')
        else:
            last_reported_at_formatted = 'Nao possui denuncias'

        return [
            ip,
            f"{abuse_confidence}%",
            f"{vt_score}",
            f"{ibm_score}" if ibm_score is not None else "",
            domain,
            country,
            city,
            last_reported_at_formatted,
            abuseipdb_link,
            virustotal_link,
            ibm_link
        ]
    except Exception as e:
        return [f"Erro ao formatar a saída para {ip}: {e}", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"]
        
def build_ip_result(ip, abuseipdb_result, virustotal_result, ibm_score,
                    city, country, domain,
                    estado_abuse=FONTE_OK, estado_vt=FONTE_OK, estado_ibm=FONTE_OK):
    # Score None = fonte nao respondeu. Nunca 0, que se confunde com "sem denuncias".
    abuse_confidence = (safe_get(abuseipdb_result, 'data', 'abuseConfidenceScore', default=0)
                        if estado_abuse == FONTE_OK else None)
    vt_score = (safe_get(virustotal_result, 'data', 'attributes',
                         'last_analysis_stats', 'malicious', default=0)
                if estado_vt == FONTE_OK else None)
    whitelisted = is_whitelisted_abuseipdb(abuseipdb_result) if estado_abuse == FONTE_OK else False

    ibm_numeric = 0
    if estado_ibm == FONTE_OK:
        try:
            ibm_numeric = float(ibm_score)
        except (ValueError, TypeError):
            ibm_numeric = 0

    has_bad_reputation = (abuse_confidence or 0) > 0 or (vt_score or 0) > 0 or ibm_numeric > 1

    fontes_indisponiveis = [nome for nome, estado in (
        ("AbuseIPDB", estado_abuse), ("VirusTotal", estado_vt), ("IBM X-Force", estado_ibm)
    ) if estado in ESTADOS_SEM_RESPOSTA]
    cota_estourada = FONTE_COTA in (estado_abuse, estado_vt, estado_ibm)

    last_reported_at = safe_get(abuseipdb_result, 'data', 'lastReportedAt')
    if last_reported_at:
        utc_time = datetime.fromisoformat(last_reported_at.replace('Z', '+00:00'))
        last_reported_at = (utc_time - timedelta(hours=3)).strftime('%Y-%m-%d %H:%M:%S')
    else:
        last_reported_at = None

    return {
        "ip": ip,
        "abuse_score": abuse_confidence,
        "vt_score": vt_score,
        "ibm_score": ibm_score,
        "whitelisted": whitelisted,
        "fontes_indisponiveis": fontes_indisponiveis,
        "cota_estourada": cota_estourada,
        "estados": {"abuse": estado_abuse, "vt": estado_vt, "ibm": estado_ibm},
        "status": (
            # Deteccao de uma fonte que respondeu ja e conclusao valida, mesmo com
            # outra fora do ar. "Limpo", nao: exige que todas tenham respondido.
            # Whitelist no AbuseIPDB nao anula deteccao das demais bases.
            "whitelisted_bad" if whitelisted and has_bad_reputation else
            "bad" if has_bad_reputation else
            "incompleto" if fontes_indisponiveis else
            "whitelisted" if whitelisted else
            "clean"
        ),
        "domain": domain,
        "country": country,
        "city": city,
        "last_report": last_reported_at,
        "links": {
            "abuse": f"https://www.abuseipdb.com/check/{ip}",
            "vt": f"https://www.virustotal.com/gui/ip-address/{ip}",
            "ibm": f"https://exchange.xforce.ibmcloud.com/ip/{ip}" if ibm_score else None
        }
    }