"""Geracao das planilhas .xlsx entregues ao cliente."""
import os
import tkinter as tk
from tkinter import filedialog

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def escolher_diretorio(parent=None, titulo="Selecione a pasta para salvar"):
    """Reaproveita a janela existente. Criar um segundo tk.Tk() dentro de um app Tk
    ja em execucao gera comportamento erratico nos dialogos e vaza a raiz nova."""
    dono = parent or tk._default_root
    temporaria = None
    if dono is None:
        temporaria = tk.Tk()
        temporaria.withdraw()
        dono = temporaria
    try:
        diretorio = filedialog.askdirectory(title=titulo, parent=dono)
    finally:
        if temporaria is not None:
            temporaria.destroy()
    return diretorio if diretorio else os.getcwd()


# O veredito ja chega com o simbolo na frente (● limpo, ▲ revisar, ✖ malicioso,
# ○ indisponivel), entao a cor sai dele sem esta camada depender do idioma.
# (fundo da linha, cor do texto do veredito) -- fundo None mantem o zebrado.
DESTAQUE_VEREDITO = {
    "✖": ("FFD9D9", "9C0006"),
    "▲": ("FFF2CC", "9C6500"),
    "●": (None, "1E7B34"),
    "○": (None, "7F7F7F"),
}


def _formatar_planilha(ws, coluna_veredito=None):
    fonte_cabecalho = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    fundo_cabecalho = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fonte_celula = Font(name="Consolas", size=10)
    alinhamento_celula = Alignment(horizontal="center", vertical="center", wrap_text=False)
    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    zebra_par = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    zebra_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for cell in ws[1]:
        cell.font = fonte_cabecalho
        cell.fill = fundo_cabecalho
        cell.alignment = alinhamento_cabecalho
        cell.border = borda
    destaques = {simbolo: PatternFill(start_color=fundo, end_color=fundo, fill_type="solid")
                 for simbolo, (fundo, _cor) in DESTAQUE_VEREDITO.items() if fundo}
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
        simbolo = _simbolo_veredito(ws, idx, coluna_veredito)
        fill = destaques.get(simbolo) or (zebra_par if idx % 2 == 0 else zebra_impar)
        for cell in row:
            cell.font = fonte_celula
            cell.alignment = alinhamento_celula
            cell.border = borda
            cell.fill = fill
        if simbolo in DESTAQUE_VEREDITO:
            cor = DESTAQUE_VEREDITO[simbolo][1]
            ws.cell(row=idx, column=coluna_veredito).font = Font(
                name="Consolas", size=10, bold=True, color=cor)
    for col in range(1, ws.max_column + 1):
        largura = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    largura = max(largura, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(largura + 4, 60)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _simbolo_veredito(ws, linha, coluna_veredito):
    if not coluna_veredito:
        return None
    valor = ws.cell(row=linha, column=coluna_veredito).value
    return str(valor).strip()[:1] if valor else None


def _preencher(ws, headers, linhas, coluna_veredito=None):
    ws.append(headers)
    for linha in linhas:
        ws.append([str(v) if v is not None else "" for v in linha])
    _formatar_planilha(ws, coluna_veredito)


# Teto do Excel. Passar dele nao levanta erro: o openpyxl so emite um UserWarning e salva
# um arquivo que o Excel pode se recusar a abrir -- quebra silenciosa, na mao do cliente.
LIMITE_NOME_ABA = 31


def _nome_aba_ips(prefixo, dominio):
    """O corte depende do prefixo: traduzi-lo muda quanto sobra para o dominio."""
    nome = dominio[:LIMITE_NOME_ABA - len(prefixo)]
    for char in ["/", "\\", "*", "?", ":", "[", "]"]:
        nome = nome.replace(char, "_")
    return prefixo + nome


# Os nomes de aba chegam prontos de quem chama: esta camada nao importa o i18n, mesmo
# motivo pelo qual a cor da linha sai do simbolo do veredito e nao do texto traduzido.
def salvar_planilha(results, headers, filename="results.xlsx", parent=None, titulo=None,
                    coluna_veredito=None, aba="Resultados"):
    caminho = os.path.join(escolher_diretorio(parent, titulo or "Selecione a pasta para salvar"),
                           filename)
    wb = Workbook()
    ws = wb.active
    ws.title = aba
    _preencher(ws, headers, results, coluna_veredito)
    wb.save(caminho)
    return caminho


def salvar_planilha_dominios(domain_results, domain_headers, ip_results_by_domain, ip_headers,
                             filename="domain_results.xlsx", parent=None, titulo=None,
                             coluna_veredito=None, aba="Dominios", prefixo_aba_ips="IPs - "):
    caminho = os.path.join(escolher_diretorio(parent, titulo or "Selecione a pasta para salvar"),
                           filename)
    wb = Workbook()
    ws = wb.active
    ws.title = aba
    _preencher(ws, domain_headers, domain_results, coluna_veredito)
    for dominio, linhas in ip_results_by_domain.items():
        _preencher(wb.create_sheet(title=_nome_aba_ips(prefixo_aba_ips, dominio)),
                   ip_headers, linhas, coluna_veredito)
    wb.save(caminho)
    return caminho
